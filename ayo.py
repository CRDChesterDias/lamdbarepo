import openpyxl

def normalize(text):
    return str(text).strip().lower()[:111] if text else ""

def match_em_to_ii(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb["ABCAEDEA"]

    headers = [cell.value for cell in ws[1]]
    col_index = {header: idx for idx, header in enumerate(headers)}

    if "Source" not in col_index or "Short description" not in col_index:
        print("❌ Required columns missing.")
        return

    ii_rows = []
    em_lookup = {}

    # Split EM and II rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        source = str(row[col_index["Source"]].value).strip().upper()
        short_desc = normalize(row[col_index["Short description"]].value)
        row_values = [cell.value for cell in row]

        if source == "EM":
            em_lookup[short_desc] = row_values
        elif source == "II":
            ii_rows.append((row_values, short_desc))

    # Prepare headers for new sheet
    em_headers = ["EM_" + h for h in headers]
    new_headers = headers + em_headers

    # Create new workbook
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "II_with_EM_matches"
    new_ws.append(new_headers)

    for ii_row, short_key in ii_rows:
        matched_em = em_lookup.get(short_key)
        combined_row = ii_row + matched_em if matched_em else ii_row + [None]*len(headers)
        new_ws.append(combined_row)

    new_wb.save(output_path)
    print(f"✅ Saved matched output to: {output_path}")
